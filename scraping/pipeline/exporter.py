"""
Export leads to CSV / Excel.

Joins Company + Website + Contact data into a single flat export.

Usage:
    python -m scraping.pipeline.exporter --format csv
    python -m scraping.pipeline.exporter --format xlsx
    python -m scraping.pipeline.exporter --min-score 55
"""

import csv
from datetime import datetime
from pathlib import Path

import click
from rich.console import Console
from sqlmodel import Session, select

from scraping.config.settings import settings
from scraping.db.init_db import get_engine
from scraping.db.models import Company, Contact, Website

console = Console()


def get_leads(min_score: int = 0, max_score: int = 100) -> list[dict]:
    """Query all leads with joined data."""
    engine = get_engine()
    leads = []

    with Session(engine) as session:
        companies = session.exec(select(Company)).all()

        for company in companies:
            website = session.exec(
                select(Website).where(Website.company_id == company.id)
            ).first()

            contact = session.exec(
                select(Contact).where(
                    Contact.company_id == company.id,
                    Contact.opted_out == False,
                )
            ).first()

            overall_score = website.overall_score if website else None

            # Filter by score if assessment exists
            if overall_score is not None:
                if not (min_score <= overall_score <= max_score):
                    continue

            leads.append({
                "company_name": company.name,
                "company_name_cn": company.name_cn,
                "domain": company.domain,
                "industry": company.industry,
                "source": company.source,
                "source_url": company.source_url,
                "country_hq": company.country_hq,
                "eu_countries": company.eu_countries_active,
                "company_size": company.company_size,
                "marketplace_url": company.marketplace_url,
                "has_standalone_site": company.has_standalone_site,
                # Website scores
                "website_url": website.url if website else "",
                "overall_score": overall_score,
                "translation_score": website.translation_score if website else None,
                "performance_score": website.performance_score if website else None,
                "seo_score": website.seo_score if website else None,
                "assessed_at": str(website.assessed_at) if website and website.assessed_at else "",
                # Contact
                "contact_email": contact.email if contact else "",
                "contact_name": contact.name if contact else "",
                "contact_title": contact.title if contact else "",
                "email_source": contact.source if contact else "",
                "email_verified": contact.email_verified if contact else False,
                "wechat_id": contact.wechat_id if contact else "",
                "contact_phone": contact.phone if contact else "",
                "linkedin_url": contact.linkedin_url if contact else "",
                # Notes
                "notes": company.notes,
            })

    return leads


def export_csv(min_score: int = 0, max_score: int = 100, output: str = ""):
    """Export leads to CSV file."""
    leads = get_leads(min_score=min_score, max_score=max_score)

    if not leads:
        console.print("[yellow]No leads to export.[/yellow]")
        return

    export_dir = Path(settings.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    if not output:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = str(export_dir / f"leads_{timestamp}.csv")

    with open(output, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=leads[0].keys())
        writer.writeheader()
        writer.writerows(leads)

    console.print(f"[green]Exported {len(leads)} leads to {output}[/green]")
    return output


def export_xlsx(min_score: int = 0, max_score: int = 100, output: str = ""):
    """Export leads to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        console.print("[red]openpyxl not installed. Run: pip install openpyxl[/red]")
        return

    leads = get_leads(min_score=min_score, max_score=max_score)

    if not leads:
        console.print("[yellow]No leads to export.[/yellow]")
        return

    export_dir = Path(settings.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    if not output:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = str(export_dir / f"leads_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header row
    headers = list(leads[0].keys())
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    header_font = Font(color="c8ff00", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_idx, lead in enumerate(leads, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=lead[key])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(output)
    console.print(f"[green]Exported {len(leads)} leads to {output}[/green]")
    return output


@click.command()
@click.option("--format", "fmt", type=click.Choice(["csv", "xlsx"]), default="csv")
@click.option("--min-score", default=0, help="Minimum website score to include")
@click.option("--max-score", default=100, help="Maximum website score (use to find poor sites)")
@click.option("--output", default="", help="Output file path")
def main(fmt: str, min_score: int, max_score: int, output: str):
    """Export leads to CSV or Excel."""
    if fmt == "csv":
        export_csv(min_score=min_score, max_score=max_score, output=output)
    else:
        export_xlsx(min_score=min_score, max_score=max_score, output=output)


if __name__ == "__main__":
    main()
